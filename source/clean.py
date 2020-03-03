import pandas as pd
import openpyxl
from source.functs import *

'''  Clean data and find errors (option 1) '''

def cleanData():
    # if response == 1:
    # ------------------------------------------------
    print('PROCESS INITIATED')
    file = "xlsxFiles/hol_hsog.xlsx"
    file2 = "xlsxFiles/holiday_formatted.xlsx"
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = ws.max_row
    end = 10  # number of steps in the process
    print("step 1 of {} complete".format(end))

    copyColumn(rows, ws, 'A%s', 'U%s')  # copy contents of column A to column U

    # --------------------------------
    print("step 2 of {} complete".format(end))

    # fill the SKU column. Fill empty cells with 0
    for cell in range(1, rows):
        y = cell + 1
        a = 'B%s' % y
        b = ws[a].value
        c = str(b)
        d = 'T%s' % cell
        if b is None:
            ws[d] = '0'
        else:
            ws[d] = c[:6]  # trim the data to 6 characters

    # --------------------------------
    print("step 3 of {} complete".format(end))

    moveOneUp(rows, ws, 'U%s', 'W%s')
    moveTwoUp(rows, ws, 'D%s', 'X%s')
    moveTwoUp(rows, ws, 'E%s', 'Y%s')
    moveTwoUp(rows, ws, 'F%s', 'Z%s')
    moveTwoUp(rows, ws, 'G%s', 'AA%s')
    moveTwoUp(rows, ws, 'I%s', 'AB%s')
    moveTwoUp(rows, ws, 'J%s', 'AC%s')
    moveTwoUp(rows, ws, 'K%s', 'AD%s')
    moveTwoUp(rows, ws, 'L%s', 'AE%s')
    moveTwoUp(rows, ws, 'M%s', 'AF%s')
    moveTwoUp(rows, ws, 'N%s', 'AG%s')
    moveTwoUp(rows, ws, 'O%s', 'AH%s')
    moveTwoUp(rows, ws, 'P%s', 'AI%s')
    moveTwoUp(rows, ws, 'Q%s', 'AJ%s')

    # --------------------------------
    print("step 4 of {} complete".format(end))

    # fill empty cells in the same column
    for cell in range(1, rows):
        y = cell + 1
        a = 'X%s' % cell
        b = 'X%s' % y
        c = ws[a].value
        d = ws[b].value
        if d is None or d == '':
            ws[b] = c

    # --------------------------------
    print("step 5 of {} complete".format(end))

    # remove empty space in every cell in column T
    for cell in range(1, rows):
        a = 'T%s' % cell
        b = ws[a].value
        c = str(b)
        if len(c) == 6:
            ws[a] = b[1:]

    # --------------------------------
    print("step 6 of {} complete".format(end))

    copyColumn(rows, ws, 'T%s', 'V%s')  # copy column T to column V

    # --------------------------------
    print("step 7 of {} complete".format(end))

    # remove first 5 spaces in every cell in column W with a length of 9
    for cell in range(1, rows):
        a = 'W%s' % cell
        b = ws[a].value
        c = str(b)
        if len(c) == 9:
            ws[a] = c[5:]

    # --------------------------------
    print("step 8 of {} complete".format(end))

    # convert the cell contents of column W to floats
    for cell in range(1, rows):
        a = 'W%s' % cell
        b = ws[a].value
        c = str(b)
        if b is None:
            continue
        elif len(c) in [1, 2, 3, 4, 5]:
            ws[a] = float(b)

    # --------------------------------
    print("step 9 of {} complete".format(end))

    wb.save(file2)
    df0 = pd.read_excel(file2)

    # select columns to keep and columns to drop
    header = []
    to_drop = []
    for cell in range(0, 36):
        header.append(cell)
        if cell < 20:
            to_drop.append(cell)
    df0.columns = header
    df0 = df0.drop(to_drop, axis='columns')

    # rename column headers
    header2 = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm',
               'n', 'o', 'p']
    df0.columns = header2

    # filter out all rows in column B that are not greater than 4
    df0 = df0[df0['b'].str.len().gt(4)]

    # -------------------------------
    print("step 10 of {} complete".format(end))

    writer = pd.ExcelWriter(file2, engine='xlsxwriter')
    df0.to_excel(writer, sheet_name="sheet")
    writer.save()

    # -------------------------------------------------
    print('PROCESS COMPLETE')
